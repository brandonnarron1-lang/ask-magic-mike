from __future__ import annotations

import hashlib
import json
import re
import subprocess
from pathlib import Path
from zipfile import BadZipFile, ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from PIL import Image

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[2]
OUT = REPO / "output" / "phase8"
VERIFY = OUT / "verification"
VERIFY.mkdir(parents=True, exist_ok=True)

PPTX = {
    "ASK_MAGIC_MIKE_PHASE8_EXECUTIVE_PRESENTATION.pptx": 15,
    "ASK_MAGIC_MIKE_CURRENT_SYSTEM_SALES_PRESENTATION.pptx": 25,
}
XLSX = [
    "ASK_MAGIC_MIKE_OPERATING_SCOREBOARD_CURRENT.xlsx",
    "ASK_MAGIC_MIKE_EVIDENCE_REGISTER_CURRENT.xlsx",
    "ASK_MAGIC_MIKE_BUDGET_ROI_MODEL_CURRENT.xlsx",
    "MESSAGE_TEMPLATE_REGISTRY_CURRENT.xlsx",
    "COMMUNICATION_PERMISSION_MATRIX_CURRENT.xlsx",
    "MESSAGE_SEQUENCE_ACCEPTANCE_MATRIX_CURRENT.xlsx",
    "PHASE7_BRANDON_EMAIL_ACCEPTANCE_CURRENT.xlsx",
    "AI_LEAD_INTELLIGENCE_EVAL_CURRENT.xlsx",
    "AI_COST_CONTROL_MATRIX_CURRENT.xlsx",
    "COPILOT_PERMISSION_MATRIX_CURRENT.xlsx",
    "MESSAGE_DELIVERABILITY_DASHBOARD_CURRENT.xlsx",
    "AI_USAGE_AND_COST_DASHBOARD_CURRENT.xlsx",
    "FORM_MESSAGE_SEQUENCE_MATRIX_CURRENT.xlsx",
]
PDFS = [
    "ASK_MAGIC_MIKE_PHASE8_EXECUTIVE_PRESENTATION.pdf",
    "ASK_MAGIC_MIKE_CURRENT_SYSTEM_SALES_PRESENTATION.pdf",
    "FORM3_CONSUMER_ACKNOWLEDGMENT_RELEASE_GATE.pdf",
    "ASK_MAGIC_MIKE_PHASE8_EXECUTIVE_SUMMARY.pdf",
    "BRANDON_PHASE8_REVIEW_GUIDE.pdf",
    "ASK_MAGIC_MIKE_PHASE8_SYSTEM_DIAGRAMS.pdf",
]
ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}


def cmd(*args: str) -> subprocess.CompletedProcess:
    return subprocess.run(args, check=True, text=True, capture_output=True)


def verify_pptx() -> list[dict]:
    results=[]
    ns = {"a":"http://schemas.openxmlformats.org/drawingml/2006/main"}
    for name, expected in PPTX.items():
        p=OUT/"presentations"/name
        with ZipFile(p) as z:
            bad=z.testzip()
            names=z.namelist()
            slides=sorted(n for n in names if re.fullmatch(r"ppt/slides/slide\d+\.xml",n))
            notes=sorted(n for n in names if re.fullmatch(r"ppt/notesSlides/notesSlide\d+\.xml",n))
            texts=0
            shapes=0
            connectors=0
            for slide in slides:
                raw=z.read(slide)
                root=ET.fromstring(raw)
                texts += len(root.findall(".//a:t",ns))
                shapes += raw.count(b"<p:sp>")
                connectors += raw.count(b"<p:cxnSp>") + raw.count(b"<a:ln")
            charts=len([n for n in names if n.startswith("ppt/charts/chart") and n.endswith(".xml")])
            media=len([n for n in names if n.startswith("ppt/media/")])
        pdf=OUT/"pdfs"/(p.stem+".pdf")
        pages=int(re.search(r"Pages:\s+(\d+)",cmd("pdfinfo",str(pdf)).stdout).group(1))
        result={"file":name,"zip_ok":bad is None,"slides":len(slides),"expected_slides":expected,"notes":len(notes),"editable_text_runs":texts,"editable_shapes":shapes,"connector_or_line_elements":connectors,"editable_charts":charts,"embedded_media":media,"pdf_pages":pages,"pass":bad is None and len(slides)==expected and len(notes)==expected and texts>0 and shapes>0 and charts>0 and pages==expected}
        results.append(result)
    (VERIFY/"presentation-verification.json").write_text(json.dumps(results,indent=2))
    return results


def verify_xlsx() -> list[dict]:
    results=[]
    for name in XLSX:
        p=OUT/"workbooks"/name
        with ZipFile(p) as z:
            bad=z.testzip()
        wb=load_workbook(p,data_only=False)
        formula_count=0
        freeze=0
        filters=0
        validations=0
        formula_errors=[]
        for ws in wb.worksheets:
            freeze += 1 if ws.freeze_panes else 0
            filters += 1 if ws.auto_filter.ref else 0
            validations += len(ws.data_validations.dataValidation)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type=="f" or (isinstance(cell.value,str) and cell.value.startswith("=")):
                        formula_count += 1
                        if any(e in str(cell.value) for e in ERRORS):
                            formula_errors.append(f"{ws.title}!{cell.coordinate}:{cell.value}")
        data_wb=load_workbook(p,data_only=True)
        cached_errors=[]
        for ws in data_wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value,str) and cell.value in ERRORS:
                        cached_errors.append(f"{ws.title}!{cell.coordinate}:{cell.value}")
        result={"file":name,"zip_ok":bad is None,"sheets":len(wb.sheetnames),"sheet_names":wb.sheetnames,"formula_count":formula_count,"freeze_panes":freeze,"filters":filters,"validations":validations,"formula_errors":formula_errors,"cached_errors":cached_errors,"recalculation_mode":wb.calculation.calcMode,"pass":bad is None and len(wb.sheetnames)>=3 and formula_count>0 and freeze>0 and filters>0 and not formula_errors and not cached_errors}
        results.append(result)
    (VERIFY/"workbook-verification.json").write_text(json.dumps(results,indent=2))
    return results


def verify_pdfs() -> list[dict]:
    results=[]
    for name in PDFS:
        p=OUT/"pdfs"/name
        q=cmd("qpdf","--check",str(p))
        info=cmd("pdfinfo",str(p)).stdout
        pages=int(re.search(r"Pages:\s+(\d+)",info).group(1))
        text=cmd("pdftotext",str(p),"-").stdout
        result={"file":name,"pages":pages,"searchable_characters":len(text.strip()),"qpdf_ok":"errors" not in q.stderr.lower(),"blank":len(text.strip())==0,"pass":pages>0 and len(text.strip())>50}
        results.append(result)
    (VERIFY/"pdf-verification.json").write_text(json.dumps(results,indent=2))
    return results


def verify_images() -> list[dict]:
    results=[]
    for p in sorted((OUT/"renders").glob("*.png"))+sorted((OUT/"workbook-previews").glob("*.png")):
        with Image.open(p) as im:
            result={"file":str(p.relative_to(OUT)),"width":im.width,"height":im.height,"nonempty":p.stat().st_size>1000,"pass":im.width>0 and im.height>0 and p.stat().st_size>1000}
            results.append(result)
    (VERIFY/"image-verification.json").write_text(json.dumps(results,indent=2))
    return results


def scan_output() -> dict:
    secret_patterns={
        "private_key":re.compile(r"-----BEGIN (?:RSA |EC |OPENSSH )?PRIVATE KEY-----"),
        "openai_key":re.compile(r"\bsk-[A-Za-z0-9_-]{20,}\b"),
        "database_url":re.compile(r"postgres(?:ql)?://[^\s<]+:[^\s<]+@",re.I),
        "bearer":re.compile(r"\bBearer\s+[A-Za-z0-9._-]{20,}\b",re.I),
    }
    pii_patterns={
        "known_private_email":re.compile(r"(?:dabnelly23|brandonnarron1)@gmail\.com",re.I),
        "known_phone":re.compile(r"(?:252[- .]?(?:245[- .]?4337|289[- .]?5194))"),
    }
    hits=[]
    readable_ext={".md",".txt",".json",".csv",".svg",".xml",".rels"}
    roots = [OUT, REPO / "docs" / "phase8", REPO / "tools" / "artifacts"]
    candidates = []
    for root in roots:
        for p in root.rglob("*"):
            if not p.is_file():
                continue
            if any(part in {"node_modules", ".venv", "__pycache__", "release"} for part in p.parts):
                continue
            candidates.append(p)
    for p in sorted(candidates):
        texts=[]
        if p.suffix.lower() in readable_ext:
            try: texts=[p.read_text(errors="ignore")]
            except OSError: pass
        elif p.suffix.lower() in {".pptx",".xlsx"}:
            try:
                with ZipFile(p) as z:
                    texts=[z.read(n).decode("utf-8","ignore") for n in z.namelist() if n.endswith((".xml",".rels"))]
            except BadZipFile: pass
        elif p.suffix.lower()==".pdf":
            try: texts=[cmd("pdftotext",str(p),"-").stdout]
            except Exception: pass
        joined="\n".join(texts)
        for label,pattern in {**secret_patterns,**pii_patterns}.items():
            if pattern.search(joined):
                hits.append({"file":str(p.relative_to(REPO)),"pattern":label})
    result={"files_scanned":len(candidates),"roots":[str(root.relative_to(REPO)) for root in roots],"hits":hits,"pass":not hits}
    (VERIFY/"secret-pii-scan.json").write_text(json.dumps(result,indent=2))
    return result


def write_report(pres, books, pdfs, images, scan):
    ok=all(x["pass"] for group in (pres,books,pdfs,images) for x in group) and scan["pass"]
    lines=[
        "# Phase 8 artifact verification report","",f"Overall: **{'PASS' if ok else 'FAIL'}**.","",
        "## Presentations","",
        "| File | Slides | Notes | Text runs | Shapes | Charts | PDF pages | Result |","|---|---:|---:|---:|---:|---:|---:|---|",
        *[f"| {x['file']} | {x['slides']} | {x['notes']} | {x['editable_text_runs']} | {x['editable_shapes']} | {x['editable_charts']} | {x['pdf_pages']} | {'PASS' if x['pass'] else 'FAIL'} |" for x in pres],"",
        "## Workbooks","","| File | Sheets | Formulas | Filters | Freeze panes | Validations | Formula errors | Result |","|---|---:|---:|---:|---:|---:|---:|---|",
        *[f"| {x['file']} | {x['sheets']} | {x['formula_count']} | {x['filters']} | {x['freeze_panes']} | {x['validations']} | {len(x['formula_errors'])+len(x['cached_errors'])} | {'PASS' if x['pass'] else 'FAIL'} |" for x in books],"",
        "## PDFs","","| File | Pages | Searchable characters | Result |","|---|---:|---:|---|",
        *[f"| {x['file']} | {x['pages']} | {x['searchable_characters']} | {'PASS' if x['pass'] else 'FAIL'} |" for x in pdfs],"",
        "## Rendering and scans","",f"- Preview/montage images verified: {len(images)}.",f"- Secret/PII files scanned: {scan['files_scanned']}.",f"- Secret/PII hits: {len(scan['hits'])}.","- Visual clipping/overflow is confirmed separately in the montage inspection and public visual-acceptance report.",""
    ]
    (REPO/"docs/phase8/ARTIFACT_VERIFICATION_REPORT_PHASE8.md").write_text("\n".join(lines))
    return ok


if __name__ == "__main__":
    pres=verify_pptx()
    books=verify_xlsx()
    pdfs=verify_pdfs()
    images=verify_images()
    scan=scan_output()
    ok=write_report(pres,books,pdfs,images,scan)
    print(json.dumps({"presentations":len(pres),"workbooks":len(books),"pdfs":len(pdfs),"images":len(images),"scan_hits":len(scan['hits']),"pass":ok}))
    raise SystemExit(0 if ok else 1)
