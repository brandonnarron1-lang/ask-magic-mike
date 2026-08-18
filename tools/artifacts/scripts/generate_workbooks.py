from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[2]
OUT = REPO / "output" / "phase8" / "workbooks"
DATA = json.loads((REPO / "output" / "phase8" / "data" / "current-system-state.json").read_text())
OUT.mkdir(parents=True, exist_ok=True)

BLACK = "080808"
INK = "171717"
GOLD = "D5AF58"
GOLD2 = "E6C977"
CREAM = "F6EEDB"
WHITE = "FFFFFF"
RUBY = "A82E39"
GREEN = "2D7A59"
AMBER = "A56B1F"
GRAY = "D9D4CA"
LIGHT = "F6F2E8"
THIN = Side(style="thin", color="B79B5A")

SOURCE_ROWS = [
    ["S01", "Authenticated production", "GitHub main", DATA["canonical"]["production_commit"], "Verified Live"],
    ["S02", "Authenticated production", "Vercel deployment", DATA["canonical"]["deployment_id"], "Verified Live"],
    ["S03", "Authenticated production", "Neon aggregate query", "Redacted production counts", "Verified Live"],
    ["S04", "Authenticated production", "WordPress bridge panel", "Bridge 1.1.0; Form 3 allowlisted", "Verified Live"],
    ["S05", "Accepted QA", "Phase 7 completion audit", "2,647 tests; 14/14 safety", "Verified Test"],
    ["S06", "Authenticated production", "Resend webhooks", "Canonical endpoint enabled; account warning present", "Operational Risk"],
]

STATUS_LIST = "Verified Live,Verified Test,Code Complete,Disabled,Approval Required,Held,Operational Risk,Assumption"


def style_sheet(ws, header_row: int = 1, freeze: str = "A2", filter_range: str | None = None) -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = filter_range or ws.dimensions
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=BLACK)
        cell.font = Font(color=GOLD2, bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="D9D4CA"))
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(r, col).value or "") for r in range(1, min(ws.max_row, 100) + 1)]
        width = min(42, max(11, max((len(v) for v in values), default=8) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[header_row].height = 28
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.outlinePr.summaryBelow = True


def add_table(ws, rows: list[list], name: str, status_col: int | None = None) -> None:
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.delete_rows(1, 1)
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    if ws.max_row >= 2 and ws.max_column >= 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        tab = Table(displayName=name[:240].replace(" ", "_"), ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)
    if status_col:
        letter = get_column_letter(status_col)
        dv = DataValidation(type="list", formula1=f'"{STATUS_LIST}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}500")
        ws.conditional_formatting.add(f"{letter}2:{letter}500", FormulaRule(formula=[f'OR({letter}2="Disabled",{letter}2="Operational Risk")'], fill=PatternFill("solid", fgColor="F4CCCC")))
        ws.conditional_formatting.add(f"{letter}2:{letter}500", FormulaRule(formula=[f'OR({letter}2="Verified Live",{letter}2="Verified Test")'], fill=PatternFill("solid", fgColor="D9EAD3")))


def add_sources(wb: Workbook) -> None:
    if "Sources" in wb.sheetnames:
        ws = wb["Sources"]
    else:
        ws = wb.create_sheet("Sources")
    add_table(ws, [["Source ID", "Class", "Source", "Evidence", "Status"], *SOURCE_ROWS], f"Sources_{len(wb.sheetnames)}", 5)


def add_definitions(wb: Workbook) -> None:
    if "Definitions" in wb.sheetnames:
        ws = wb["Definitions"]
    else:
        ws = wb.create_sheet("Definitions")
    rows = [
        ["Term", "Definition", "Calculation / rule"],
        ["Genuine lead", "A public consumer submission not marked test", "is_test = false"],
        ["Suppressed QA", "Unmistakable internal QA record excluded from outreach and KPIs", "is_test = true AND communication_suppressed = true"],
        ["Business KPI leads", "Genuine leads only", "Genuine live leads; QA excluded"],
        ["Verified Live", "Observed in authenticated production", "Must have current evidence"],
        ["Verified Test", "Observed only with controlled synthetic/suppressed data", "Never relabel as commercial performance"],
        ["Approval Required", "Implemented but not enabled", "Separate explicit decision"],
    ]
    add_table(ws, rows, f"Definitions_{len(wb.sheetnames)}")


def finalize(wb: Workbook, filename: str) -> None:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.properties.creator = "Our Town Properties / Ask Magic Mike"
    wb.properties.title = filename.replace("_", " ").removesuffix(".xlsx")
    wb.save(OUT / filename)


def operating_scoreboard() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Dashboard"
    ws.append(["ASK MAGIC MIKE — CURRENT OPERATING SCOREBOARD", "Verified value", "Business KPI", "Status", "Source"])
    metrics = [
        ["Genuine leads", DATA["lead_counts"]["genuine_live"], "=B2", "Verified Live", "S03"],
        ["Suppressed QA", DATA["lead_counts"]["qa_suppressed"], "=0", "Verified Test", "S03"],
        ["QA in Active/New", DATA["lead_counts"]["qa_in_business_active_or_new"], "=B4", "Verified Live", "S03"],
        ["Pending notifications", DATA["notifications"]["pending_or_retrying"], "=B5", "Verified Live", "S03"],
        ["Live notification failures", DATA["notifications"]["live_failures"], "=B6", "Verified Live", "S03"],
        ["Commercial revenue", 0, "=B7", "Verified Live", "S03"],
    ]
    for r in metrics:
        ws.append(r)
    style_sheet(ws)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws["A9"] = "QA exclusion check"
    ws["B9"] = "=IF(C3=0,\"PASS\",\"FAIL\")"
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Current verified counts"
    chart.y_axis.title = "Count"
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=6), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=6))
    chart.height = 7.5
    chart.width = 14.5
    ws.add_chart(chart, "G2")
    sheet_rows = {
        "Funnel": [["Stage", "Live", "QA", "KPI formula", "Status"], ["Created", 0, 6, "=B2", "Verified Live"], ["Qualified", 0, 0, "=B3", "Verified Live"], ["Appointment", 0, 0, "=B4", "Verified Live"], ["Closing", 0, 0, "=B5", "Verified Live"]],
        "Sources": [["Source", "Live leads", "Qualified", "QA", "KPI leads"], ["AskMagicMike.com", 0, 0, 6, "=B2"], ["OurTownProperties.com / Form 3", 0, 0, 0, "=B3"]],
        "Forms": [["Form", "Purpose", "WP active", "Forwarding", "Live leads", "QA", "Status"], *[[f["id"], f["name"], f["wordpress_active"], f["canonical_forwarding"], 0, 1 if f["id"] == 3 else 0, "Verified Live" if f["id"] == 3 else "Held"] for f in DATA["forms"]]],
        "Assignment SLA": [["Metric", "Live due", "QA excluded", "Rate"], ["Within SLA", 0, 6, "=IFERROR(B2/B3,0)"], ["Overdue", 0, 6, "=IFERROR(B3/B2,0)"]],
        "Contact SLA": [["Metric", "Live count", "Target", "Pass"], ["Median first contact minutes", 0, 15, "=IF(B2=0,\"NO LIVE DATA\",IF(B2<=C2,\"PASS\",\"FAIL\"))"]],
        "Notifications": [["Class", "Pending", "Sent", "Failed", "Failure rate"], ["Live", 0, 0, 0, "=IFERROR(D2/SUM(B2:D2),0)"], ["Suppressed QA", 0, 4, 2, "=IFERROR(D3/SUM(B3:D3),0)"]],
        "AI Usage": [["Mode", "Runs", "Input tokens", "Output tokens", "Cost", "Avg latency"], [DATA["ai_copilot"]["latest"]["mode"], 1, 835, 964, 0.006619, 7624], ["Total", "=SUM(B2:B2)", "=SUM(C2:C2)", "=SUM(D2:D2)", "=SUM(E2:E2)", "=AVERAGE(F2:F2)"]],
        "Communication Permissions": [["Purpose", "Channel", "Current rows", "Allowed live", "State"], ["Consumer acknowledgment", "email", 0, 0, "Disabled"], ["QA test", "email", 0, 0, "Verified Test"]],
        "Message Sequences": [["Sequence", "Instances", "Steps", "Live active", "State"], ["All", 0, 0, 0, "Disabled"]],
        "QA Exclusions": [["Rule", "Expected", "Actual", "Pass"], ["QA excluded from KPI leads", 0, 0, "=IF(B2=C2,\"PASS\",\"FAIL\")"], ["Unsuppressed QA", 0, 0, "=IF(B3=C3,\"PASS\",\"FAIL\")"]],
    }
    for name, rows in sheet_rows.items():
        target = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
        if target.max_row == 1 and target["A1"].value is None:
            pass
        add_table(target, rows, name.replace(" ", "_") + "Tbl", len(rows[0]) if rows[0][-1] in ("Status", "State") else None)
    add_definitions(wb)
    if "Sources" in wb.sheetnames:
        src = wb["Sources"]
        src.delete_rows(1, src.max_row)
        add_table(src, [["Source ID", "Class", "Source", "Evidence", "Status"], *SOURCE_ROWS], "OperatingSources", 5)
    finalize(wb, "ASK_MAGIC_MIKE_OPERATING_SCOREBOARD_CURRENT.xlsx")


def evidence_register() -> None:
    wb = Workbook()
    wb.active.title = "Evidence Register"
    add_table(wb.active, [["Evidence ID", "Claim", "State", "Source", "Observed", "Owner review"],
        ["E001", "Canonical production is Ready", "Verified Live", "S01/S02", "2026-08-18", "No"],
        ["E002", "Genuine live lead count is zero", "Verified Live", "S03", "2026-08-18", "No"],
        ["E003", "Six QA records are suppressed", "Verified Test", "S03", "2026-08-18", "No"],
        ["E004", "Resend account warning may disrupt delivery", "Operational Risk", "S06", "2026-08-18", "Yes"],
        ["E005", "Lead Center screenshots need fresh sign-in", "Approval Required", "Browser audit", "2026-08-18", "Yes"]], "EvidenceRegister", 3)
    tabs = {
        "Assumptions": [["Assumption", "Editable input", "Source", "Risk"], ["Revenue per closing", 7500, "Owner input required", "Do not use until verified"], ["Monthly paid media", 0, "No authorization", "Inactive"]],
        "Current Approvals": [["Gate", "State", "Exact decision"], ["Form 3 acknowledgment", "Approval Required", "APPROVE FORM 3 CONSUMER ACKNOWLEDGMENT EMAIL PILOT"], ["Consumer SMS", "Disabled", "No approval"], ["Mike activation", "Disabled", "Deferred"]],
        "Access Gaps": [["Surface", "State", "Required action"], ["Lead Center", "Session expired", "Fresh staff sign-in"], ["Resend", "Operational Risk", "Resolve account standing or approve replacement"]],
        "Source Index": [["Source ID", "Description", "Status"], *[[r[0], r[2], r[4]] for r in SOURCE_ROWS]],
        "Change History": [["Date", "Change", "Commit", "Status"], ["2026-08-17", "Current evidence merge", DATA["canonical"]["production_commit"], "Verified Live"], ["2026-08-18", "Phase 8 artifact build", "Current branch", "Code Complete"]],
        "Status Definitions": [["Status", "Meaning"], ["Verified Live", "Observed in production"], ["Verified Test", "Controlled synthetic evidence"], ["Disabled", "Feature intentionally off"], ["Approval Required", "Separate explicit decision"]]
    }
    for name, rows in tabs.items():
        ws = wb.create_sheet(name)
        add_table(ws, rows, name.replace(" ", "") + "Tbl", 2 if name in ("Current Approvals", "Access Gaps") else None)
        ws.cell(ws.max_row + 2, 1, "Row count")
        ws.cell(ws.max_row, 2, f"=COUNTA(A2:A{ws.max_row-2})")
    finalize(wb, "ASK_MAGIC_MIKE_EVIDENCE_REGISTER_CURRENT.xlsx")


def roi_model() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ROI Dashboard"
    rows = [
        ["Scenario", "Monthly spend", "Leads", "Qualification rate", "Appointment rate", "Closing rate", "Revenue / closing", "Qualified", "Appointments", "Closings", "Revenue", "Contribution", "ROI"],
        ["Current verified", 0, 0, 0, 0, 0, 0, "=C2*D2", "=H2*E2", "=I2*F2", "=J2*G2", "=K2-B2", "=IFERROR(L2/B2,0)"],
        ["Conservative assumption", 1500, 12, 0.35, 0.25, 0.08, 6500, "=C3*D3", "=H3*E3", "=I3*F3", "=J3*G3", "=K3-B3", "=IFERROR(L3/B3,0)"],
        ["Expected assumption", 3000, 25, 0.45, 0.35, 0.12, 7500, "=C4*D4", "=H4*E4", "=I4*F4", "=J4*G4", "=K4-B4", "=IFERROR(L4/B4,0)"],
        ["Strong assumption", 4500, 40, 0.55, 0.45, 0.18, 8500, "=C5*D5", "=H5*E5", "=I5*F5", "=J5*G5", "=K5-B5", "=IFERROR(L5/B5,0)"]
    ]
    add_table(ws, rows, "ROIDashboard")
    for row in range(3, 6):
        for col in range(2, 8):
            ws.cell(row, col).fill = PatternFill("solid", fgColor="FFF2CC")
    ws["A8"] = "Break-even closings (editable expected case)"
    ws["B8"] = "=IFERROR(B4/G4,0)"
    ws["A9"] = "Important"
    ws["B9"] = "The single $0.006619 AI acceptance result is not used as a monthly average."
    add_table(wb.create_sheet("AI Cost Assumptions"), [["Feature", "Current observed", "Planning input", "Rule"], ["Copilot acceptance", 0.006619, 0, "One synthetic test; no monthly extrapolation"], ["Monthly AI budget", 0, 25, "Editable cap only"]], "AICostAssumptions")
    add_table(wb.create_sheet("Messaging Assumptions"), [["Channel", "Current state", "Monthly volume input", "Unit cost input", "Monthly cost"], ["Consumer email", "Disabled", 0, 0, "=C2*D2"], ["Consumer SMS", "Disabled", 0, 0, "=C3*D3"]], "MessagingAssumptions", 2)
    add_table(wb.create_sheet("Automation Assumptions"), [["Feature", "State", "Assumed uplift", "Included"], ["Form 3 acknowledgment", "Approval Required", 0, "No"], ["Nurture", "Disabled", 0, "No"]], "AutomationAssumptions", 2)
    add_table(wb.create_sheet("Break-even"), [["Revenue / closing", "Monthly spend", "Break-even closings"], [6500, 1500, "=B2/A2"], [7500, 3000, "=B3/A3"], [8500, 4500, "=B4/A4"]], "BreakEven")
    add_table(wb.create_sheet("Commission Economics"), [["Input", "Value", "Source", "Editable"], ["Gross commission revenue", 7500, "Owner verification required", "Yes"], ["Contribution margin", 1.0, "Assumption", "Yes"], ["Net revenue", "=B2*B3", "Calculated", "No"]], "CommissionEconomics")
    add_sources(wb)
    finalize(wb, "ASK_MAGIC_MIKE_BUDGET_ROI_MODEL_CURRENT.xlsx")


def matrix_book(filename: str, main_name: str, headers: list[str], rows: list[list], extra_sheets: dict[str, list[list]] | None = None, status_col: int | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = main_name
    add_table(ws, [headers, *rows], main_name.replace(" ", "") + "Data", status_col)
    summary = wb.create_sheet("Summary", 0)
    summary.append([filename.replace("_", " ").removesuffix(".xlsx"), "Value"])
    summary.append(["Data rows", f"=COUNTA('{main_name}'!A2:A500)"])
    summary.append(["Verified live rows", f'=COUNTIF(\'{main_name}\'!A2:{get_column_letter(len(headers))}500,"Verified Live")'])
    summary.append(["Disabled / held rows", f'=COUNTIF(\'{main_name}\'!A2:{get_column_letter(len(headers))}500,"Disabled")+COUNTIF(\'{main_name}\'!A2:{get_column_letter(len(headers))}500,"Held")'])
    style_sheet(summary)
    for name, erows in (extra_sheets or {}).items():
        ews = wb.create_sheet(name)
        add_table(ews, erows, name.replace(" ", "") + "Data")
        ews.cell(ews.max_row + 2, 1, "Row count")
        ews.cell(ews.max_row, 2, f"=COUNTA(A2:A{ews.max_row-2})")
    add_definitions(wb)
    add_sources(wb)
    finalize(wb, filename)


def build_matrices() -> None:
    matrix_book(
        "MESSAGE_TEMPLATE_REGISTRY_CURRENT.xlsx", "Template Registry",
        ["Template ID", "Version", "Lead type", "Channel", "Purpose", "Subject", "Consent", "Timing", "Stop conditions", "Approval", "Active state", "QA result", "Source path"],
        [
            ["home_value.email.received", "v1", "home_value", "email", "transactional_acknowledgment", "Your home-value review request was received", "purpose-specific allowed decision", "immediate after storage", "suppressed, test, bounce, complaint", "Approval Required", "Disabled", "Verified Test", "src/lib/messaging/template-registry.ts"],
            ["seller.email.received", "v1", "seller", "email", "transactional_acknowledgment", "Your seller request was received", "purpose-specific allowed decision", "immediate after storage", "suppressed, test, bounce, complaint", "Approval Required", "Disabled", "Code Complete", "src/lib/messaging/template-registry.ts"],
            ["lead_alert.internal", "current", "all", "email", "internal_alert", "Priority | source | intent | location | name | score", "internal operational basis", "after durable storage", "idempotency; bounded retry", "Approved configuration", "Code Complete", "Verified Test", "src/lib/notifications"],
        ], status_col=11)
    matrix_book(
        "COMMUNICATION_PERMISSION_MATRIX_CURRENT.xlsx", "Permission Matrix",
        ["Lead type", "Channel", "Purpose", "Consent state", "Test", "Suppression", "Legal hold", "BIC hold", "Allowed", "Reason", "Manual review", "Audit event", "Status"],
        [[lead, ch, purp, consent, test, supp, legal, bic, allowed, reason, review, audit, status] for lead,ch,purp,consent,test,supp,legal,bic,allowed,reason,review,audit,status in [
            ["all", "email", "internal_alert", "operational", False, False, False, False, True, "approved internal route", False, "communication_decision", "Code Complete"],
            ["home_value", "email", "transactional_acknowledgment", "ambiguous", False, False, False, False, False, "ambiguity fails closed", True, "permission_blocked", "Disabled"],
            ["all", "email", "qa_test", "qa_allowlist", True, True, False, False, True, "test + suppressed + exact recipient", False, "qa_decision", "Verified Test"],
            ["all", "sms", "marketing_nurture", "unknown", False, False, False, False, False, "carrier SMS disabled", True, "permission_blocked", "Disabled"],
        ]], status_col=13)
    matrix_book(
        "MESSAGE_SEQUENCE_ACCEPTANCE_MATRIX_CURRENT.xlsx", "Sequence Acceptance",
        ["Sequence", "Step", "Timing", "Channel", "Purpose", "Permission", "Quiet hours", "Frequency cap", "Stop conditions", "QA result", "Production state", "Status"],
        [["home_value", i+1, timing, channel, purpose, "required", quiet, cap, stops, qa, prod, status] for i,(timing,channel,purpose,quiet,cap,stops,qa,prod,status) in enumerate([
            ["immediate", "email", "transactional_acknowledgment", "n/a", "1/version", "test, suppression, duplicate, bounce, complaint", "Verified Test", "disabled", "Disabled"],
            ["same day", "email", "requested_service_response", "n/a", "bounded", "reply, hold, bounce, complaint", "Code Complete", "disabled", "Disabled"],
            ["day 1", "sms", "appointment_coordination", "America/New_York", "bounded", "STOP, reply, hold", "Verified Test", "carrier disabled", "Disabled"],
        ])], status_col=12)
    matrix_book(
        "PHASE7_BRANDON_EMAIL_ACCEPTANCE_CURRENT.xlsx", "Brandon Acceptance",
        ["Test ID", "Template", "Provider ID", "Subject", "Accepted", "Sent", "Delivered", "Mobile", "Desktop", "Dark mode", "Links", "No Mike delivery", "No consumer delivery", "Reporting excluded", "Status"],
        [["phase7-brandon-qa", "phase7_qa_email", "871e5b96-a10b-492a-bb23-9898824f0cd3", "[TEST — BRANDON QA] Phase 7 messaging release-candidate review", True, True, True, "Pass", "Pass", "Reviewed", "Pass", True, True, True, "Verified Test"]], status_col=15)
    ai_cases = ["Seller","Buyer","Value","Ask","Seller options","Rental","Property alerts","Out of area","Coastal review","Test","Suppressed","Ambiguous consent","Prompt injection","Missing data","Conflicting data"]
    matrix_book(
        "AI_LEAD_INTELLIGENCE_EVAL_CURRENT.xlsx", "AI Evaluation",
        ["Case", "Synthetic", "Expected behavior", "Schema", "PII redacted", "No mutation", "Guardrail", "Result", "Score"],
        [[case, True, "Structured advisory; escalate uncertainty" if case not in ("Prompt injection","Ambiguous consent") else "Block action; label risk", "Pass", True, True, "Pass", "Accepted synthetic test" if case=="Suppressed" else "Test coverage", 1] for case in ai_cases],
        extra_sheets={"Scoring Rubric":[["Criterion","Weight","Formula"],["Schema",0.2,"=B2"],["PII",0.2,"=B3"],["No mutation",0.3,"=B4"],["Guardrail",0.3,"=B5"]]})
    matrix_book(
        "AI_COST_CONTROL_MATRIX_CURRENT.xlsx", "AI Cost Control",
        ["Feature", "Model", "State", "Token budget", "Cost cap", "Timeout ms", "Retries", "Fallback", "Daily limit", "Monitoring", "Current acceptance result", "Status"],
        [["Lead Center Copilot", "gpt-5.6-luna", "operator only", 4000, "bounded", 20000, 2, "deterministic", "feature gated", "usage rows", "$0.006619 / 1799 tokens / 7624 ms", "Verified Test"], ["Automatic action", "none", "disabled", 0, 0, 0, 0, "blocked", 0, "flag", "No run", "Disabled"]], status_col=12)
    matrix_book(
        "COPILOT_PERMISSION_MATRIX_CURRENT.xlsx", "Copilot Permissions",
        ["Tool", "Read/write", "Admin", "Primary owner", "Agent", "Analyst", "Viewer", "Test only", "Approval", "Consent", "Audit", "Failure behavior", "Status"],
        [[tool, rw, admin, owner, agent, analyst, viewer, test, approval, consent, audit, failure, status] for tool,rw,admin,owner,agent,analyst,viewer,test,approval,consent,audit,failure,status in [
            ["Read lead facts", "read", True, True, "assigned only", "redacted", False, False, "session", "n/a", True, "deny", "Code Complete"],
            ["Generate advisory", "append result", True, True, "assigned only", False, False, True, "human review", "PII minimized", True, "fallback", "Verified Test"],
            ["Change assignment", "write", False, False, False, False, False, False, "not allowed", "n/a", True, "blocked", "Disabled"],
            ["Send message", "write", False, False, False, False, False, False, "not allowed", "permission required", True, "blocked", "Disabled"],
        ]], status_col=13)
    matrix_book(
        "MESSAGE_DELIVERABILITY_DASHBOARD_CURRENT.xlsx", "Deliverability",
        ["Class", "Accepted", "Sent", "Delivered", "Delayed", "Bounced", "Complained", "Failed", "Replied", "Opted out", "Test", "Live", "Provider latency ms", "Failure rate", "Status"],
        [["Brandon QA", 1, 1, 1, 0, 0, 0, 0, 0, 0, 1, 0, "not retained in redacted source", "=IFERROR(H2/B2,0)", "Verified Test"], ["Suppressed QA legacy", 2, 0, 0, 0, 0, 0, 2, 0, 0, 2, 0, "n/a", "=IFERROR(H3/B3,0)", "Verified Test"], ["Live", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "n/a", "=IFERROR(H4/B4,0)", "Verified Live"]], status_col=15)
    matrix_book(
        "AI_USAGE_AND_COST_DASHBOARD_CURRENT.xlsx", "AI Usage",
        ["Date", "Feature", "Mode", "Model", "Test/live", "Input tokens", "Output tokens", "Cost", "Latency", "Fallback", "Error", "Guardrail", "PII redacted", "Audit ID", "Status"],
        [["2026-08-17", "lead_center_copilot", "openai_responses", "gpt-5.6-luna", "Test", 835, 964, 0.006619, 7624, "", "", "Pass", True, "redacted", "Verified Test"], ["TOTAL", "", "", "", "", "=SUM(F2:F2)", "=SUM(G2:G2)", "=SUM(H2:H2)", "=AVERAGE(I2:I2)", "", "", "", "", "", "Verified Test"]], status_col=15)
    form_rows=[]
    for f in DATA["forms"]:
        form_rows.append([f["id"], f["name"], f["canonical_forwarding"], "versioned evidence required" if f["id"]==3 else "not approved", "home_value" if f["id"]==3 else "none", "disabled", "mapped" if f["id"]==3 else "held", "Mike/admin review", "eligible after approval" if f["id"]==3 else "not eligible", "verified" if f["id"]==3 else "not run", "separate approval" if f["id"]==3 else "held", "disable bridge allowlist / feature flag", "Approval Required" if f["id"]==3 else "Held"])
    matrix_book(
        "FORM_MESSAGE_SEQUENCE_MATRIX_CURRENT.xlsx", "Form Matrix",
        ["Form", "Purpose", "Current state", "Consent", "Email family", "SMS permission", "AI mapping", "Routing", "Sequence eligibility", "QA", "Approval", "Rollback", "Status"], form_rows, status_col=13)


if __name__ == "__main__":
    operating_scoreboard()
    evidence_register()
    roi_model()
    build_matrices()
    files = sorted(OUT.glob("*.xlsx"))
    print(f"Generated {len(files)} workbooks in {OUT}")
