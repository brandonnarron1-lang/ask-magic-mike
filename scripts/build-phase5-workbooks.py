#!/usr/bin/env python3
"""Build Phase 5 operational workbooks without secrets or genuine customer PII."""

from __future__ import annotations

import csv
import importlib.util
import os
import re
from pathlib import Path
from urllib.parse import urlencode

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "output" / "phase5" / "spreadsheets"
OUT.mkdir(parents=True, exist_ok=True)
os.environ["AMM_WORKBOOK_OUTPUT_DIR"] = str(OUT)

spec = importlib.util.spec_from_file_location("phase2_workbooks", ROOT / "scripts" / "build-phase2-workbooks.py")
phase2 = importlib.util.module_from_spec(spec)
assert spec and spec.loader
spec.loader.exec_module(phase2)

GOLD = phase2.GOLD
GOLD_LIGHT = phase2.GOLD_LIGHT
BLACK = phase2.BLACK
CHARCOAL = phase2.CHARCOAL
WHITE = phase2.WHITE
GREEN = phase2.GREEN
GREEN_LIGHT = phase2.GREEN_LIGHT
AMBER = phase2.AMBER
AMBER_LIGHT = phase2.AMBER_LIGHT
RED = phase2.RED
RED_LIGHT = phase2.RED_LIGHT
INPUT = phase2.INPUT
THIN = phase2.THIN


def save(wb: Workbook, name: str) -> Path:
    path = OUT / name
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        if ws.max_row and ws.max_column:
            ws.print_area = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(path)
    check = load_workbook(path, data_only=False, read_only=True)
    assert check.sheetnames
    check.close()
    return path


def input_validation(ws, cell_range: str, values: str) -> None:
    dv = DataValidation(type="list", formula1=f'"{values}"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def build_first_live_acceptance() -> Path:
    wb, summary = phase2.base_book(
        "First Genuine Lead Acceptance Template",
        "Private operator checklist. Do not place customer PII in screenshots, exported evidence, or this shared template.",
    )
    summary["A7"] = "Current genuine live leads"
    summary["B7"] = 0
    summary["A8"] = "Current suppressed QA"
    summary["B8"] = 6
    summary["A9"] = "Current QA in Active/New"
    summary["B9"] = 0
    summary["A10"] = "Acceptance readiness"
    summary["B10"] = "='Acceptance Checklist'!F27"
    summary["A11"] = "Evidence classification"
    summary["B11"] = "VERIFIED baseline; first-live outcome remains NOT YET OBSERVED"
    phase2.style_key_value(summary, "A7:B11")

    checklist = wb.create_sheet("Acceptance Checklist")
    headers = ["Step", "Control", "Expected evidence", "Owner", "Status", "Safe evidence reference"]
    controls = [
        (1, "Not QA", "test=false with no explicit QA evidence", "Administrator"),
        (2, "Not suppressed", "communication_suppressed=false", "Administrator"),
        (3, "Durable persistence", "One Neon lead ID exists before notifications", "System"),
        (4, "Exactly one canonical lead", "Idempotency/dedupe query returns one master", "System"),
        (5, "Source and attribution", "Source URL, placement, first/last touch, UTMs/click IDs", "Administrator"),
        (6, "Consent evidence", "Exact text/version/timestamp and permitted channels", "Administrator"),
        (7, "Explainable score", "Score, grade, factors, and weights visible", "System"),
        (8, "Approved owner", "Mike or explicit approved fallback with routing reason", "System"),
        (9, "Assignment audit", "Prior/new owner, actor, reason, timestamp", "System"),
        (10, "Internal email outbox", "One idempotent notification record", "System"),
        (11, "Provider result", "Provider ID and terminal/known status", "System"),
        (12, "Hidden audit copy", "BCC result confirmed without revealing address", "Administrator"),
        (13, "Web Push", "Attempt/result only when enrolled device exists", "System"),
        (14, "No duplicate", "No duplicate lead or duplicate alert", "System"),
        (15, "Assignment timer", "Internal timer started from durable creation", "System"),
        (16, "First-contact timer", "Timer started; no public response-time promise", "System"),
        (17, "Escalation", "Unassigned/undelivered condition escalates", "System"),
        (18, "Human acknowledgment", "Authorized operator records acknowledgment", "Assigned owner"),
        (19, "Contact outcome", "Permitted outcome recorded without excess PII", "Assigned owner"),
        (20, "Channel guardrails", "No carrier SMS or unsupported marketing", "Administrator"),
        (21, "Reporting inclusion", "Included only in live KPIs after genuine classification", "System"),
        (22, "Private reconciliation", "Sensitive details remain in authenticated Lead Center", "Administrator"),
    ]
    rows = [[step, control, evidence, owner, "PENDING", ""] for step, control, evidence, owner in controls]
    phase2.title_sheet(checklist, "First Genuine Lead Acceptance", "Complete only after a genuine public submission. Never fabricate a prospect.", len(headers))
    phase2.write_table(checklist, headers, rows, widths=[8,30,62,22,16,48])
    input_validation(checklist, "E5:E26", "PENDING,PASS,FAIL,N/A")
    checklist["E27"] = "Overall"
    checklist["F27"] = '=IF(COUNTIF(E5:E26,"FAIL")>0,"FAIL",IF(COUNTIF(E5:E26,"PENDING")>0,"PENDING","PASS"))'
    phase2.style_key_value(checklist, "E27:F27")

    timing = wb.create_sheet("Timing Ledger")
    phase2.title_sheet(timing, "Private Timing Ledger", "Use lead ID/correlation ID only; do not duplicate contact details here.", 8)
    phase2.write_table(timing, ["Lead ID", "Created UTC", "Assigned UTC", "Acknowledged UTC", "First contact UTC", "Assignment sec", "Contact sec", "Result"], [], widths=[39,24,24,26,26,18,18,22])
    for row in range(5, 105):
        timing.cell(row, 6, f'=IF(OR(B{row}="",C{row}=""),"",(C{row}-B{row})*86400)')
        timing.cell(row, 7, f'=IF(OR(B{row}="",E{row}=""),"",(E{row}-B{row})*86400)')
        timing.cell(row, 8, f'=IF(A{row}="","",IF(OR(F{row}>120,G{row}>300),"REVIEW","WITHIN INTERNAL TARGET"))')
        for col in range(1, 9):
            timing.cell(row, col).border = Border(bottom=THIN)
    timing.conditional_formatting.add("F5:G104", CellIsRule(operator="greaterThan", formula=["300"], fill=PatternFill("solid", fgColor=RED_LIGHT)))

    evidence = wb.create_sheet("Evidence Register")
    phase2.title_sheet(evidence, "Acceptance Evidence Register", "Reference secure records; never paste reset links, Push endpoints, credentials, BCC values, or customer messages.", 6)
    phase2.write_table(evidence, ["Control", "Evidence type", "Safe reference", "Verified by", "Verified at", "Status"], [], widths=[34,28,56,24,24,16])
    input_validation(evidence, "F5:F104", "PENDING,VERIFIED,FAILED,NOT APPLICABLE")
    return save(wb, "FIRST_LIVE_LEAD_ACCEPTANCE_TEMPLATE.xlsx")


def read_csv(name: str) -> list[dict[str, str]]:
    with (ROOT / "docs" / name).open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def build_utm_library() -> Path:
    source = read_csv("ASK_MAGIC_MIKE_UTM_LINK_LIBRARY.csv")
    wb, summary = phase2.base_book("Ask Magic Mike UTM Link Library", "Approved-source link register. Tagged URLs are formula-derived and contain no personal data.")
    summary["A7"] = "Placements"
    summary["B7"] = "=COUNTA('Tagged Links'!A5:A104)"
    summary["A8"] = "Publication-approved"
    summary["B8"] = '=COUNTIF(\'Tagged Links\'!M5:M104,"PUBLISHED")'
    summary["A9"] = "Paid media active"
    summary["B9"] = 0
    phase2.style_key_value(summary, "A7:B9")
    ws = wb.create_sheet("Tagged Links")
    headers = ["Placement", "Base URL", "utm_source", "utm_medium", "utm_campaign", "utm_content", "Lead type", "Owner", "Tagged URL", "QR asset", "HTTP check", "QR scan", "Publication status"]
    rows = []
    for row in source:
        query = urlencode({
            "utm_source": row["UTM source"],
            "utm_medium": row["UTM medium"],
            "utm_campaign": "owned_traffic_phase5",
            "utm_content": row["UTM content"],
        })
        tagged = f'{row["Base URL"]}?{query}'
        qr_name = "qr-" + "-".join(part for part in re.sub(r"[^a-z0-9]+", "-", row["Placement"].lower()).split("-") if part) + ".png"
        rows.append([row["Placement"], row["Base URL"], row["UTM source"], row["UTM medium"], "owned_traffic_phase5", row["UTM content"], row["Intended lead type"], row["Owner"], tagged, qr_name, "PASS", "PASS", "STAGED - APPROVAL REQUIRED"])
    phase2.title_sheet(ws, "Tagged Links", "Do not append lead names, email addresses, phone numbers, or property details to URLs.", len(headers))
    phase2.write_table(ws, headers, rows, widths=[27,44,26,16,28,30,22,16,86,30,18,18,30])
    for row in range(5, 5 + len(rows)):
        ws.cell(row, 9).hyperlink = ws.cell(row, 9).value
        ws.cell(row, 9).style = "Hyperlink"
    input_validation(ws, f"K5:M{4+len(rows)}", "PENDING,PASS,FAIL,STAGED - APPROVAL REQUIRED,PUBLISHED")

    conventions = wb.create_sheet("Conventions")
    phase2.title_sheet(conventions, "UTM Conventions", "Stable lower-case names support first-touch and last-touch attribution.", 4)
    phase2.write_table(conventions, ["Field", "Rule", "Example", "Do not use"], [
        ["utm_source", "Origin domain/platform", "ourtownproperties.com", "Names or campaign prose"],
        ["utm_medium", "Traffic mechanism", "referral, email, social, qr", "PII"],
        ["utm_campaign", "Stable initiative", "owned_traffic_phase5", "Dates unless intentionally versioned"],
        ["utm_content", "Exact placement", "home_value_page", "Property/customer details"],
        ["click IDs", "Capture approved IDs server-side", "gclid, fbclid", "Expose in analytics with PII"],
    ], widths=[24,54,34,46])
    return save(wb, "ASK_MAGIC_MIKE_UTM_LINK_LIBRARY.xlsx")


def build_content_queue() -> Path:
    source = read_csv("ZERO_SPEND_CONTENT_APPROVAL_QUEUE.csv")
    wb, summary = phase2.base_book("Zero-Spend Content Approval Queue", "Thirty-day owned-content queue. Nothing in this workbook is authorization to publish.")
    summary["A7"] = "Draft assets"
    summary["B7"] = "=COUNTA('Approval Queue'!A5:A34)"
    summary["A8"] = "Published externally"
    summary["B8"] = '=COUNTIF(\'Approval Queue\'!G5:G34,"PUBLISHED")'
    summary["A9"] = "Paid media"
    summary["B9"] = "DISABLED"
    phase2.style_key_value(summary, "A7:B9")
    ws = wb.create_sheet("Approval Queue")
    headers = list(source[0].keys()) + ["Approver", "Approval date", "Published URL", "Performance note"]
    rows = [[row[h] for h in source[0].keys()] + ["", "", "", ""] for row in source]
    phase2.title_sheet(ws, "Approval Queue", "BIC/content/legal decisions remain human approvals. Publication status defaults to draft.", len(headers))
    phase2.write_table(ws, headers, rows, widths=[8,48,25,22,24,30,23,28,28,28,45,55,24,18,54,46])
    input_validation(ws, "G5:G34", "DRAFT,DRAFT - FUTURE,APPROVED,PUBLISHED,REJECTED")
    return save(wb, "ZERO_SPEND_CONTENT_APPROVAL_QUEUE.xlsx")


def build_day7_report() -> Path:
    wb, summary = phase2.base_book("Ask Magic Mike Seven-Day Operations Report", "Formula-driven operating review. Test and suppressed rows are excluded from every live KPI.")
    metrics = [
        ("Genuine leads", '=SUMIFS(\'Daily Metrics\'!D5:D11,\'Daily Metrics\'!B5:B11,"LIVE",\'Daily Metrics\'!C5:C11,"NOT SUPPRESSED")'),
        ("Qualified leads", "=SUM('Daily Metrics'!E5:E11)"),
        ("Contacted leads", "=SUM('Daily Metrics'!F5:F11)"),
        ("Appointments", "=SUM('Daily Metrics'!G5:G11)"),
        ("Signed clients", "=SUM('Daily Metrics'!H5:H11)"),
        ("Active canonical forms", "=MAX('Daily Metrics'!I5:I11)"),
        ("Average assignment seconds", '=IFERROR(AVERAGEIF(\'Daily Metrics\'!D5:D11,">0",\'Daily Metrics\'!K5:K11),0)'),
        ("Average first-contact seconds", '=IFERROR(AVERAGEIF(\'Daily Metrics\'!D5:D11,">0",\'Daily Metrics\'!L5:L11),0)'),
        ("Notification failures", "=SUM('Daily Metrics'!M5:M11)"),
        ("Duplicate rate", '=IFERROR(SUM(\'Daily Metrics\'!N5:N11)/B7,0)'),
        ("Unassigned rate", '=IFERROR(SUM(\'Daily Metrics\'!O5:O11)/B7,0)'),
        ("SLA breaches", "=SUM('Daily Metrics'!P5:P11)"),
        ("Published owned assets", "=SUM('Daily Metrics'!Q5:Q11)"),
    ]
    for idx, (label, formula) in enumerate(metrics, 7):
        summary.cell(idx, 1, label)
        summary.cell(idx, 2, formula)
    phase2.style_key_value(summary, "A7:B19")
    summary["B16"].number_format = "0.0%"
    summary["B17"].number_format = "0.0%"
    ws = wb.create_sheet("Daily Metrics")
    headers = [
        "Day", "Classification", "Suppression", "Genuine leads", "Qualified",
        "Contacted", "Appointments", "Signed clients", "Active canonical forms",
        "Top source", "Average assignment seconds", "Average first-contact seconds",
        "Notification failures", "Duplicates", "Unassigned", "SLA breaches",
        "Published owned assets", "Notes",
    ]
    rows = [
        [f"Day {day}", "LIVE", "NOT SUPPRESSED", 0, 0, 0, 0, 0, 1, "", 0, 0, 0, 0, 0, 0, 0, ""]
        for day in range(1, 8)
    ]
    phase2.title_sheet(ws, "Daily Metrics", "Enter actual observed outcomes only; no forecasts in actual columns.", len(headers))
    phase2.write_table(
        ws,
        headers,
        rows,
        widths=[12,20,22,18,16,16,18,18,24,28,28,30,24,16,18,18,28,55],
    )
    input_validation(ws, "B5:B11", "LIVE,QA")
    input_validation(ws, "C5:C11", "NOT SUPPRESSED,SUPPRESSED")
    for row in range(5, 12):
        for col in range(2, 18):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
    chart = BarChart()
    chart.title = "Seven-Day Operational Outcomes"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Day"
    data = Reference(ws, min_col=4, max_col=8, min_row=4, max_row=11)
    cats = Reference(ws, min_col=1, min_row=5, max_row=11)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, "K4")

    review = wb.create_sheet("Operating Review")
    phase2.title_sheet(review, "Seven-Day Review", "Use evidence-backed answers; distinguish verified facts, owner decisions, and hypotheses.", 5)
    phase2.write_table(review, ["Question", "Answer", "Evidence", "Owner", "Status"], [
        ["Did every genuine lead persist before notification?", "", "", "Administrator", "PENDING"],
        ["Were assignment and first-contact targets met?", "", "", "Administrator", "PENDING"],
        ["Did any QA/test record enter live KPIs?", "", "", "System owner", "PENDING"],
        ["Did any email fail or duplicate?", "", "", "System owner", "PENDING"],
        ["Which owned placement produced qualified intent?", "", "", "Marketing", "PENDING"],
        ["What single change is justified next?", "", "", "Owner/BIC", "PENDING"],
    ], widths=[55,62,48,24,16])
    input_validation(review, "E5:E10", "PENDING,PASS,FAIL,NOT OBSERVED")
    return save(wb, "ASK_MAGIC_MIKE_7_DAY_OPERATIONS_REPORT.xlsx")


def build_incident_matrix() -> Path:
    source = read_csv("INCIDENT_ESCALATION_MATRIX.csv")
    wb, summary = phase2.base_book("Incident Escalation Matrix", "Operational escalation reference. Targets are internal controls, not consumer promises.")
    summary["A7"] = "Critical controls"
    summary["B7"] = '=COUNTIF(\'Escalation Matrix\'!A5:A104,"Critical")'
    summary["A8"] = "High controls"
    summary["B8"] = '=COUNTIF(\'Escalation Matrix\'!A5:A104,"High")'
    summary["A9"] = "Open incidents"
    summary["B9"] = '=COUNTIF(\'Incident Log\'!H5:H104,"OPEN")'
    phase2.style_key_value(summary, "A7:B9")
    ws = wb.create_sheet("Escalation Matrix")
    headers = list(source[0].keys())
    rows = [[row[h] for h in headers] for row in source]
    phase2.title_sheet(ws, "Escalation Matrix", "Do not include credentials, raw Push endpoints, private BCC values, or customer contact details.", len(headers))
    phase2.write_table(ws, headers, rows, widths=[16,42,36,55,24,26,44,45])
    log = wb.create_sheet("Incident Log")
    phase2.title_sheet(log, "Incident Log", "Use correlation IDs and safe references only.", 9)
    phase2.write_table(log, ["Incident ID", "Opened", "Severity", "Condition", "Safe reference", "Owner", "Action", "Status", "Closed"], [], widths=[24,24,16,45,45,24,55,18,24])
    input_validation(log, "C5:C104", "Critical,High,Warning")
    input_validation(log, "H5:H104", "OPEN,MONITORING,RESOLVED")
    return save(wb, "INCIDENT_ESCALATION_MATRIX.xlsx")


def build_day1_report() -> Path:
    wb, summary = phase2.base_book("Ask Magic Mike Day 1 Operations Report", "Editable first-day ledger. Actual live customer details remain inside the authenticated Lead Center.")
    summary["A7"] = "Production health"
    summary["B7"] = "PASS - live and ready"
    summary["A8"] = "Genuine leads"
    summary["B8"] = 0
    summary["A9"] = "Suppressed QA"
    summary["B9"] = 6
    summary["A10"] = "QA in Active/New"
    summary["B10"] = 0
    summary["A11"] = "Active canonical forms"
    summary["B11"] = 1
    summary["A12"] = "Active Push devices"
    summary["B12"] = 0
    phase2.style_key_value(summary, "A7:B12")
    ops = wb.create_sheet("Opening and Closing")
    phase2.title_sheet(ops, "Day 1 Opening and Closing", "Record only safe status and evidence references.", 6)
    phase2.write_table(ops, ["Control", "Opening result", "Closing result", "Owner", "Safe evidence", "Notes"], [
        ["Public health and routes", "PASS", "", "System owner", "Health monitor", ""],
        ["Active/New contains zero QA", "PASS", "", "Administrator", "Lead Center active view", ""],
        ["Notification queue clear", "PASS", "", "System owner", "Neon aggregate", ""],
        ["Form 3 bridge healthy", "PASS", "", "Administrator", "WordPress bridge health", ""],
        ["First-live monitor active", "PASS", "", "System owner", "Vercel cron logs", ""],
        ["Carrier SMS disabled", "PASS", "", "System owner", "Configuration audit", ""],
        ["No unapproved content published", "PASS", "", "Owner", "Approval queue", ""],
    ], widths=[45,22,22,24,45,50])
    return save(wb, "ASK_MAGIC_MIKE_DAY1_OPERATIONS_REPORT.xlsx")


def validate_outputs(paths: list[Path]) -> None:
    for path in paths:
        workbook = load_workbook(path, data_only=False, read_only=True)
        assert workbook.sheetnames, path
        formulas = 0
        for ws in workbook.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas += 1
        workbook.close()
        if path.name in {"FIRST_LIVE_LEAD_ACCEPTANCE_TEMPLATE.xlsx", "ASK_MAGIC_MIKE_7_DAY_OPERATIONS_REPORT.xlsx"}:
            assert formulas > 0, path


def main() -> None:
    paths = [Path(builder()) for builder in phase2.builders]
    paths.extend([
        build_first_live_acceptance(),
        build_utm_library(),
        build_content_queue(),
        build_day1_report(),
        build_day7_report(),
        build_incident_matrix(),
    ])
    validate_outputs(paths)
    print("\n".join(str(path) for path in paths))


if __name__ == "__main__":
    main()
