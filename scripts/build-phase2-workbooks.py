#!/usr/bin/env python3
"""Build Ask Magic Mike Phase 2 operating workbooks with no secrets or live PII."""

from __future__ import annotations

from datetime import datetime
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT = Path(os.environ.get("AMM_WORKBOOK_OUTPUT_DIR", ROOT / "output" / "phase2" / "spreadsheets"))
OUT.mkdir(parents=True, exist_ok=True)

GOLD = "D9A441"
GOLD_LIGHT = "F6E8BF"
BLACK = "111111"
CHARCOAL = "242424"
WHITE = "FFFFFF"
GREEN = "2E7D32"
GREEN_LIGHT = "E2F0D9"
AMBER = "B26A00"
AMBER_LIGHT = "FFF2CC"
RED = "B3261E"
RED_LIGHT = "F4CCCC"
BLUE_LIGHT = "DDEBF7"
GRAY = "E7E6E6"
INPUT = "FFF2CC"
THIN = Side(style="thin", color="D9D9D9")


def base_book(title: str, subtitle: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=GOLD)
    ws["A1"].fill = PatternFill("solid", fgColor=BLACK)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A2:H2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Aptos", size=10, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30
    ws["A4"] = "Generated"
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M ET")
    ws["A5"] = "Data classification"
    ws["B5"] = "Operational metadata only - no secrets, private BCC, or live customer PII"
    style_key_value(ws, "A4:B5")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 72
    configure_print(ws, landscape=False)
    return wb, ws


def configure_print(ws, landscape: bool = True):
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.print_options.horizontalCentered = True


def style_key_value(ws, cell_range: str):
    for row in ws[cell_range]:
        row[0].font = Font(bold=True, color=CHARCOAL)
        row[0].fill = PatternFill("solid", fgColor=GOLD_LIGHT)
        for cell in row:
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def title_sheet(ws, title: str, note: str, end_col: int):
    configure_print(ws)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.print_title_rows = "1:4"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    c = ws.cell(1, 1, title)
    c.font = Font(name="Aptos Display", size=18, bold=True, color=GOLD)
    c.fill = PatternFill("solid", fgColor=BLACK)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    c = ws.cell(2, 1, note)
    c.font = Font(size=9, color="666666")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 28


def write_table(ws, headers, rows, start_row=4, widths=None, table_name=None):
    for col, header in enumerate(headers, 1):
        c = ws.cell(start_row, col, header)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=CHARCOAL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[start_row].height = 34
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, value in enumerate(row, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = Border(bottom=THIN)
            if r_idx % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F8F8F8")
        ws.row_dimensions[r_idx].height = 34
    ws.auto_filter.ref = f"A{start_row}:{chr(64 + min(len(headers), 26))}{start_row + len(rows)}" if len(headers) <= 26 else None
    if widths:
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    return start_row + len(rows)


def add_status_rules(ws, status_col: str, start: int, end: int):
    rng = f"{status_col}{start}:{status_col}{end}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("ACTIVE",{status_col}{start}))'], fill=PatternFill("solid", fgColor=GREEN_LIGHT), font=Font(color=GREEN, bold=True)))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("BLOCK",{status_col}{start})),ISNUMBER(SEARCH("FAIL",{status_col}{start})),ISNUMBER(SEARCH("DEFER",{status_col}{start})))'], fill=PatternFill("solid", fgColor=RED_LIGHT), font=Font(color=RED, bold=True)))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("SHADOW",{status_col}{start})),ISNUMBER(SEARCH("PENDING",{status_col}{start})))'], fill=PatternFill("solid", fgColor=AMBER_LIGHT), font=Font(color=AMBER, bold=True)))


def save(wb, name: str):
    path = OUT / name
    wb.save(path)
    # Structural reopen proves the package is a valid XLSX.
    check = load_workbook(path, data_only=False, read_only=True)
    assert check.sheetnames
    check.close()
    return path


forms = [
    [1, "Contact Us", "https://www.ourtownproperties.com/contact-us/", "general_question", "SHADOW - READY FOR QA", "1.1.0", "No", "Yes", "POST /api/leads", "Admin Notification (active)", "Yes after canonical QA", "No native Consent field; absent permissions default false", "UTM/click IDs via bridge context", "name 1; email 2; phone 6; subject 3; message 4", "Route to Mike", "Mike/admin review", "Not run", "", "Per-form 1.1.0 config + GF export", "Use denied consent defaults until approved wording"],
    [2, "Cash Offer Form", "No rendered sitemap placement found", "seller_cash_offer", "SHADOW - MAPPING REQUIRED", "1.1.0", "No", "Yes", "POST /api/leads", "Admin Notification (active)", "Yes after canonical QA", "No native Consent field", "UTM/click IDs via bridge context", "name 1; phone 7; email 2; address 6", "Route to Mike", "Mike", "Not run", "", "Per-form 1.1.0 config + GF export", "Brokerage review of seller-options copy required"],
    [3, "Home Value Form", "https://www.ourtownproperties.com/how-much-is-your-home-worth/", "home_value", "ACTIVE - CANONICAL", "1.1.0", "Yes", "Yes", "POST /api/leads", "Admin Notification (inactive duplicate)", "Corrected", "Denied call/email/SMS stored for QA", "Verified first/last touch and UTMs", "address 6; name 1; email 2; phone 7", "Route to Mike", "Mike", "PASS", "2026-08-14", "wordpress-canonical-bridge-1.0.0-rollback.zip", "Entry 1549; one canonical lead; one alert; replay idempotent"],
    [4, "Join Our Team Form", "https://www.ourtownproperties.com/join-our-team/", "recruiting", "NOT A LEAD FORM", "1.1.0", "No", "Yes", "Not activated", "Admin Notification (active)", "No", "No native Consent field", "UTM context available", "name 1; phone 7; email 2; license 8; address 6", "Recruiting/admin only", "Unverified", "Not applicable", "", "GF export", "Keep outside consumer routing"],
    [5, "Rental Property Search", "No rendered sitemap placement found", "renter", "SHADOW - MAPPING REQUIRED", "1.1.0", "No", "Yes", "POST /api/leads", "Admin Notification (active)", "Yes after canonical QA", "No native Consent field", "UTM/click IDs via bridge context", "name 1; phone 7; email 2; address 6; area 8", "Mike or admin review", "Unverified", "Not run", "", "Per-form 1.1.0 config + GF export", "Confirm public placement and recipient"],
    [6, "Short Term Home Rentals form", "https://www.ourtownproperties.com/short-term-home-rentals/", "renter", "SHADOW - READY FOR QA", "1.1.0", "No", "Yes", "POST /api/leads", "Admin Notification (active)", "Yes after canonical QA", "No native Consent field", "UTM/click IDs via bridge context", "name 1; phone 7; email 2; details 8", "Mike/admin review", "Mike/admin review", "Not run", "", "Per-form 1.1.0 config + GF export", "Controlled QA after denied-consent mapping"],
    [7, "Never miss a property!", "Global/sitewide placement", "buyer_property_alert", "DEFERRED - LEGAL REVIEW", "1.1.0", "No", "Yes", "POST /api/leads after consent fix", "Admin Notification (active)", "Yes after canonical QA", "No explicit marketing Consent field", "Page and campaign context required", "name 1; phone 7; email 2; message 8", "Mike until approved mapping", "Mike/admin review", "Blocked", "", "Per-form 1.1.0 config + GF export", "Possible genuine entry 1550 preserved; do not use as QA"],
]


def build_activation_matrix():
    wb, summary = base_book("WordPress Form Activation Matrix", "Controlled form-by-form canonical activation register. Form 3 is the verified baseline.")
    summary["A7"] = "Active canonical forms"; summary["B7"] = "='Form Inventory'!F12"
    summary["A8"] = "Shadow forms"; summary["B8"] = "='Form Inventory'!F13"
    summary["A9"] = "Blocked/deferred forms"; summary["B9"] = "='Form Inventory'!F14"
    summary["A10"] = "Live prospects in Neon"; summary["B10"] = 0
    summary["A11"] = "Unsuppressed tests"; summary["B11"] = 0
    style_key_value(summary, "A7:B11")

    ws = wb.create_sheet("Form Inventory")
    headers = ["Form ID","Form name","Public URL","Lead type","Current status","Bridge version","Allowlisted?","Local Gravity entry?","Canonical endpoint","Current notifications","Duplicate-notification candidate","Consent fields","Attribution fields","Canonical field mapping","Routing rule","Assigned owner","QA status","Activation date","Rollback package","Notes"]
    title_sheet(ws, "Form Inventory", "Source: authenticated Gravity Forms and rendered public sitemap, 2026-08-14.", len(headers))
    write_table(ws, headers, forms, widths=[9,25,42,20,28,14,13,18,22,31,20,35,30,43,24,20,18,16,34,55])
    ws["E12"] = "Summary"
    ws["F12"] = '=COUNTIF(E5:E11,"ACTIVE*")'
    ws["F13"] = '=COUNTIF(E5:E11,"SHADOW*")'
    ws["F14"] = '=COUNTIF(E5:E11,"DEFERRED*")+COUNTIF(E5:E11,"NOT A LEAD FORM")'
    add_status_rules(ws, "E", 5, 11)

    gates = ["Backup complete","Field mapping complete","Consent mapping complete","Attribution mapping complete","Routing tested","Signature tested","Replay tested","Local entry verified","Neon record verified","Canonical ID verified","Internal alert verified","Hidden BCC verified","Consumer message suppressed","SMS suppressed","Duplicate notification corrected","Rollback tested","Production health passed"]
    ws = wb.create_sheet("Activation Gates")
    headers = ["Form ID","Form name"] + gates + ["Completion %","Activation eligibility","Remaining blockers"]
    title_sheet(ws, "Activation Gates", "Use PASS, FAIL, N/A, or PENDING. Eligibility is formula driven.", len(headers))
    rows=[]
    for f in forms:
        values = [f[0],f[1]]
        for gate in gates:
            if f[0] == 3: values.append("PASS")
            elif f[0] == 4: values.append("N/A")
            else: values.append("PENDING")
        rows.append(values)
    write_table(ws, headers, rows, widths=[9,28]+[14]*len(gates)+[14,22,44])
    for r in range(5,12):
        start_col=3; end_col=2+len(gates)
        applicable = f'COUNTIF(C{r}:{ws.cell(r,end_col).column_letter}{r},"PASS")+COUNTIF(C{r}:{ws.cell(r,end_col).column_letter}{r},"FAIL")+COUNTIF(C{r}:{ws.cell(r,end_col).column_letter}{r},"PENDING")'
        ws.cell(r,end_col+1, f'=IFERROR(COUNTIF(C{r}:{ws.cell(r,end_col).column_letter}{r},"PASS")/({applicable}),0)')
        ws.cell(r,end_col+1).number_format="0%"
        ws.cell(r,end_col+2, f'=IF(({applicable})=0,"NOT APPLICABLE",IF(COUNTIF(C{r}:{ws.cell(r,end_col).column_letter}{r},"FAIL")>0,"BLOCKED",IF(COUNTIF(C{r}:{ws.cell(r,end_col).column_letter}{r},"PENDING")>0,"NOT READY","ELIGIBLE")))')
        ws.cell(r,end_col+3, f'=IF(({applicable})=0,"No applicable consumer lead gates",COUNTIF(C{r}:{ws.cell(r,end_col).column_letter}{r},"PENDING")&" pending; "&COUNTIF(C{r}:{ws.cell(r,end_col).column_letter}{r},"FAIL")&" failed")')
    dv=DataValidation(type="list",formula1='"PASS,FAIL,N/A,PENDING"',allow_blank=False); ws.add_data_validation(dv); dv.add(f"C5:{ws.cell(11,2+len(gates)).coordinate}")

    ws = wb.create_sheet("QA Evidence")
    headers=["QA test ID","Form ID","Date","Gravity entry ID","Canonical lead ID","Resend provider ID","Notification count","Replay result","Suppression result","Test-record status","Cleanup/reconciliation status","Evidence path","Pass or fail"]
    rows=[["FORM3-QA-20260814",3,"2026-08-14",1549,"70f63f35-2478-4738-b84c-bc1a89b8482c","bf31a582-e4a3-45cb-a7f1-5cb89121626f",1,"Same lead; no new message","PASS","is_test=true","Suppressed and retained","docs/FORM3_PRODUCTION_RECONCILIATION.md","PASS"]]
    title_sheet(ws,"QA Evidence","No private BCC values or secrets.",len(headers)); write_table(ws,headers,rows,widths=[24,9,14,18,39,39,18,26,22,22,30,45,14])

    ws = wb.create_sheet("Rollback Register")
    headers=["Form","Previous state","Current state","Notification state","Bridge state","Backup package","Rollback steps","Rollback test date","Owner","Status"]
    rows=[[f"Form {f[0]} - {f[1]}","Native notification active; not allowlisted" if f[0]!=3 else "Native notification active; bridge shadow","Canonical active" if f[0]==3 else "Unchanged/shadow",f[9],"Allowlisted" if f[0]==3 else "Not allowlisted",f[18],"Disable per-form forwarding; restore only proven duplicate notification; preserve entries","2026-08-14" if f[0]==3 else "",f[15],"PASS" if f[0]==3 else "AVAILABLE"] for f in forms]
    title_sheet(ws,"Rollback Register","Never delete Gravity entries during rollback.",len(headers)); write_table(ws,headers,rows,widths=[30,38,28,32,20,38,60,20,22,16]); add_status_rules(ws,"J",5,11)

    return save(wb,"WORDPRESS_FORM_ACTIVATION_MATRIX.xlsx")


def build_scoreboard():
    wb, ws = base_book("Ask Magic Mike Operating Scoreboard", "Editable operations tracker. Rows marked Test? = Yes are excluded from every KPI.")
    ws["A7"]="New live leads"; ws["B7"]='=COUNTIFS(\'Lead Log\'!A5:A504,"<>",\'Lead Log\'!B5:B504,"No")'
    ws["A8"]="Qualified live leads"; ws["B8"]='=COUNTIFS(\'Lead Log\'!B5:B504,"No",\'Lead Log\'!D5:D504,"Yes")'
    ws["A9"]="Appointments"; ws["B9"]='=COUNTIFS(\'Lead Log\'!B5:B504,"No",\'Lead Log\'!G5:G504,"Set")'
    ws["A10"]="Signed clients"; ws["B10"]='=COUNTIFS(\'Lead Log\'!B5:B504,"No",\'Lead Log\'!H5:H504,"Yes")'
    ws["A11"]="Gross attributed revenue"; ws["B11"]='=SUMIFS(\'Lead Log\'!R5:R504,\'Lead Log\'!B5:B504,"No")'; ws["B11"].number_format='$#,##0'
    ws["A12"]="Delivery failures"; ws["B12"]='=COUNTIFS(\'Lead Log\'!B5:B504,"No",\'Lead Log\'!P5:P504,">0")'
    style_key_value(ws,"A7:B12")
    log=wb.create_sheet("Lead Log")
    headers=["Lead ID","Test?","Created date","Qualified?","Assignment seconds","First-contact seconds","Appointment status","Signed client?","Closing status","Source","Campaign","Form","Duplicate count","Stage","Owner","Delivery failures","Cost USD","Gross attributed revenue USD","Notes"]
    title_sheet(log,"Lead Log","Paste operational outcomes here. Never include secrets. Test rows remain editable but excluded from KPIs.",len(headers))
    write_table(log,headers,[],widths=[38,10,16,13,20,22,22,18,20,24,28,24,18,20,22,20,16,28,45])
    for r in range(5,505):
        for c in range(1,len(headers)+1):
            log.cell(r,c).border=Border(bottom=THIN); log.cell(r,c).alignment=Alignment(wrap_text=True,vertical="top")
        for c in [2,4,7,8,9,14]: log.cell(r,c).fill=PatternFill("solid",fgColor=INPUT)
    for col, values in [(2,"Yes,No"),(4,"Yes,No"),(7,"Not set,Requested,Set,Completed,Cancelled"),(8,"Yes,No"),(9,"Open,Pending,Closed,Lost"),(14,"new,assigned,attempted,contacted,nurture,appointment,signed,closed,bad lead")]:
        dv=DataValidation(type="list",formula1=f'"{values}"'); log.add_data_validation(dv); dv.add(f"{log.cell(5,col).column_letter}5:{log.cell(504,col).column_letter}504")
    defs=wb.create_sheet("Metric Definitions")
    title_sheet(defs,"Metric Definitions","Operational targets are internal targets, not public guarantees.",4)
    write_table(defs,["Metric","Definition","Target","Test handling"],[["Assignment time","Seconds from durable lead creation to owner acceptance","120 seconds","Exclude tests"],["First-contact time","Seconds from durable lead creation to first recorded human contact","300 seconds","Exclude tests"],["Qualified lead","Owner marks Qualified? = Yes with an actual business reason","No fabricated target","Exclude tests"],["Attributed revenue","Gross revenue entered only after closing and linked to source/campaign","Actual only","Exclude tests"]],widths=[28,65,28,20])
    return save(wb,"ASK_MAGIC_MIKE_OPERATING_SCOREBOARD.xlsx")


def build_roster():
    wb,ws=base_book("Lead Center User Roster Template","Owner-completed roster. No invitation is sent until identity and role are verified.")
    r=wb.create_sheet("User Roster"); headers=["Approved?","Full name","Work email","Role","Territory","Lead permissions","Export permitted?","Status","Approved by","Approval date","Notes"]
    title_sheet(r,"User Roster","Yellow cells are owner inputs. Do not invent users or use personal addresses without approval.",len(headers)); write_table(r,headers,[],widths=[13,25,34,24,26,42,20,20,25,18,45])
    for row in range(5,55):
        for col in range(1,len(headers)+1): r.cell(row,col).fill=PatternFill("solid",fgColor=INPUT); r.cell(row,col).border=Border(bottom=THIN)
    for col,vals in [(1,"Yes,No"),(4,"administrator,primary_lead_owner,approved_agent,read_only_analyst"),(7,"Yes,No"),(8,"pending,approved,provisioned,revoked")]:
        dv=DataValidation(type="list",formula1=f'"{vals}"'); r.add_data_validation(dv); dv.add(f"{r.cell(5,col).column_letter}5:{r.cell(54,col).column_letter}54")
    roles=wb.create_sheet("Role Definitions"); title_sheet(roles,"Role Definitions","Least-privilege baseline.",4); write_table(roles,["Role","Lead scope","Allowed outcomes","Prohibited"],[["administrator","All authorized leads","Users, assignment, exports, audits, routing","No secret exposure"],["primary_lead_owner","Assigned and explicitly authorized leads","Contact, status, notes, appointments, closings","No user or secret administration"],["approved_agent","Assigned leads only","Status and authorized notes/outcomes","No unassigned leads or raw exports"],["read_only_analyst","Non-sensitive reporting","Attribution and performance views","No raw PII export or mutation"]],widths=[28,40,55,48])
    return save(wb,"LEAD_CENTER_USER_ROSTER_TEMPLATE.xlsx")


def build_access_matrix():
    wb,ws=base_book("Lead Center Access Matrix","Server-side role and object-level authorization contract.")
    m=wb.create_sheet("Role Permission Matrix"); headers=["Capability","Administrator","Primary Lead Owner","Approved Agent","Read-Only Analyst","Enforcement"]
    rows=[]
    capabilities=[("View all leads","Allow","Scoped","Deny","Deny"),("View assigned leads","Allow","Allow","Allow","PII redacted"),("Assign/reassign","Allow","Deny","Deny","Deny"),("Update stage","Allow","Allow","Allow assigned","Deny"),("Add notes/tasks","Allow","Allow","Allow assigned","Deny"),("Export raw PII","Allow + audit","Deny","Deny","Deny"),("View source reporting","Allow","Allow scoped","Allow scoped","Allow non-sensitive"),("Manage users","Allow","Deny","Deny","Deny"),("Configure routing","Allow","Deny","Deny","Deny"),("View audit history","Allow","Own scope","Own scope","Non-sensitive only"),("Revoke sessions","Allow","Own session","Own session","Own session")]
    for cap,a,p,g,v in capabilities: rows.append([cap,a,p,g,v,"Server route + repository query scope + audit"])
    title_sheet(m,"Role Permission Matrix","Every protected API must enforce role and resource scope server-side.",len(headers)); write_table(m,headers,rows,widths=[30,22,24,24,24,55])
    e=wb.create_sheet("Endpoint Controls"); title_sheet(e,"Endpoint Controls","Cutover stays pending until a verified administrator is provisioned.",6); write_table(e,["Surface","Authentication","Authorization","CSRF","Cache","Status"],[["/admin","Per-user session planned; Basic fallback current","Role gate","Origin + token for writes","no-store","IMPLEMENTATION PENDING"],["/admin/api/*","Session + route check","Role and object scope","Required for cookie auth","no-store","TEST REQUIRED"],["Exports","Session","Administrator only + audit","Required","no-store","TEST REQUIRED"],["Emergency access","Basic fallback retained until cutover","Administrator only","N/A","no-store","CURRENT FALLBACK"]],widths=[30,42,40,32,18,28]); add_status_rules(e,"F",5,8)
    return save(wb,"LEAD_CENTER_ACCESS_MATRIX.xlsx")


def build_push_register():
    wb,ws=base_book("Web Push Device Register","No subscription endpoint or cryptographic key may be entered in this workbook.")
    r=wb.create_sheet("Device Register"); headers=["Owner","Role","Device name","Browser","Enrollment status","Enrolled date","[TEST] push received?","Deep link verified?","Duplicate handled?","Revocation verified?","Last verification","Notes"]
    rows=[["Mike","primary","","Safari/approved browser","NOT ENROLLED","","No","No","Not tested","Not tested","","Physical permission required"],["Brandon","copy","","Safari/approved browser","NOT ENROLLED","","No","No","Not tested","Not tested","","Physical permission required"]]
    title_sheet(r,"Device Register","Physical permission cannot be bypassed. Carrier SMS stays disabled.",len(headers)); write_table(r,headers,rows,widths=[18,14,28,25,24,18,22,22,22,22,20,45]); add_status_rules(r,"E",5,6)
    a=wb.create_sheet("Acceptance Gates"); title_sheet(a,"Acceptance Gates","Complete with the device owner present.",5); write_table(a,["Gate","Mike","Brandon","Evidence","Status"],[[x,"PENDING","PENDING","No private endpoint values", "PENDING"] for x in ["Secure authentication","Permission granted","Subscription stored","[TEST] push received","Deep link secure","Duplicate enrollment handled","Revocation works"]],widths=[35,18,18,45,20])
    return save(wb,"WEB_PUSH_DEVICE_REGISTER.xlsx")


def build_monitoring_matrix():
    wb,ws=base_book("Monitoring Alert Matrix","Point-in-time checks are not continuous monitoring. Status shows what is truly scheduled.")
    m=wb.create_sheet("Alert Matrix"); headers=["Monitor","Target/condition","Check type","Frequency","Severity","Owner","Alert channel","Normal reporting","Runbook","Schedule status"]
    rows=[
        ["Liveness","/api/health/live != 200","Scheduled synthetic","Hourly","Critical","System owner","Failure-only email/Web Push when enrolled","Daily digest","PRODUCTION_MONITORING_RUNBOOK.md","ACTIVE - GITHUB SCHEDULE"],
        ["Readiness","/api/health/ready != 200","Scheduled synthetic","Hourly","Critical","System owner","Failure-only","Daily digest","PRODUCTION_MONITORING_RUNBOOK.md","ACTIVE - GITHUB SCHEDULE"],
        ["Public routes","/, /sell, /buy, /value, /ask, widget","Scheduled synthetic","Hourly","High","System owner","Failure-only","Daily digest","PRODUCTION_MONITORING_RUNBOOK.md","ACTIVE - 9/9 PASS"],
        ["Admin denial","Anonymous /admin != 401","Scheduled synthetic","Hourly","Critical","System owner","Immediate","Daily digest","SECURITY_REVIEW_PHASE5.md","ACTIVE - GITHUB SCHEDULE"],
        ["Queue failures","failed live notifications > 0","Database reconciliation","Every 2 min","Critical","System owner","Immediate","Daily digest","EMAIL_DELIVERY_SPEC.md","ACTIVE - FIRST-LIVE CRON"],
        ["Unassigned lead","Live unsuppressed lead has no owner","Database reconciliation","Every 2 min","Critical","Mike/admin","Immediate","Daily digest","FIRST_LIVE_LEAD_RESPONSE_RUNBOOK.md","ACTIVE - FIRST-LIVE CRON"],
        ["SLA breach","Acceptance >2m or contact >5m","Database reconciliation","Hourly","High","Mike/admin","Immediate","Daily digest","FIRST_LIVE_LEAD_RESPONSE_RUNBOOK.md","ACTIVE - SLA CRON"],
        ["Unsuppressed test","Any QA lead is not suppressed","Database reconciliation","Every 2 min","Critical","System owner","Immediate","Daily digest","LEAD_QUEUE_INVARIANTS.md","ACTIVE - FIRST-LIVE CRON"],
        ["Unauthorized form forward","Bridge form outside allowlist","WordPress reconciliation","Hourly","Critical","System owner","Immediate","Daily digest","WORDPRESS_INTEGRATION.md","SCRIPT REQUIRED"],
        ["NellySelly isolation","Identifier or project crossover","Release + source audit","Every release","Critical","System owner","Immediate","Daily digest","ARCHITECTURE.md","POINT-IN-TIME PASS"],
    ]
    title_sheet(m,"Alert Matrix","Do not send routine all-clear messages every few minutes.",len(headers)); write_table(m,headers,rows,widths=[28,45,28,18,16,22,24,18,42,34]); add_status_rules(m,"J",5,14)
    return save(wb,"MONITORING_ALERT_MATRIX.xlsx")


def build_attribution_matrix():
    wb,ws=base_book("Form Attribution Acceptance Matrix","Canonical field acceptance by form. QA and test traffic remain excluded from KPIs and audiences.")
    a=wb.create_sheet("Acceptance Matrix"); fields=["First touch","Latest touch","UTM source","UTM medium","UTM campaign","UTM content","UTM term","Landing page","Referrer","Form ID","Form name","Placement","Property/listing ID","gclid","fbclid","Consent","Lead type","Assignment","Appointment","Signed client","Closing","Attributed revenue","Test exclusion"]
    headers=["Field"]+[f"Form {i}" for i in range(1,8)]+["Canonical definition"]
    rows=[]
    for field in fields:
        vals=[field]
        for i in range(1,8):
            if i==3: vals.append("PASS")
            elif i==4: vals.append("N/A")
            elif i==7 and field in ["Consent","Test exclusion"]: vals.append("BLOCKED")
            else: vals.append("PENDING")
        vals.append("Stored server-side; no raw PII in analytics" if field not in ["Appointment","Signed client","Closing","Attributed revenue"] else "Recorded only after an actual business outcome")
        rows.append(vals)
    title_sheet(a,"Acceptance Matrix","Nested WordPress click-ID compatibility is covered by PR #140.",len(headers)); write_table(a,headers,rows,widths=[28]+[15]*7+[58])
    for col in range(2,9): add_status_rules(a,a.cell(4,col).column_letter,5,4+len(rows))
    return save(wb,"FORM_ATTRIBUTION_ACCEPTANCE_MATRIX.xlsx")


def build_test_register():
    wb,ws=base_book("Test Lead Register","Production test-governance register. No synthetic record is a live prospect.")
    r=wb.create_sheet("Production Tests"); headers=["Test lead ID","Purpose","Date","Environment","Form","Suppression state","Notification state","Audit state","Reconciliation state","Retention decision"]
    rows=[
        ["a1a7e899-9b2e-4ffe-968f-1e10728d60e8","Initial public production QA/master","2026-08-11","production","Ask Magic Mike public seller","Suppressed","No outbound - invalid key stage","Audited","Reconciled","Retain as QA evidence"],
        ["8609b5e2-da81-49b0-8db9-c113af6894a3","Server-side QA marker verification","2026-08-11","production","Ask Magic Mike public seller","Suppressed","No outbound","Audited","Reconciled","Retain as QA evidence"],
        ["59bba7cf-fe27-42c3-adb6-27b27727e5c7","Verified sender end-to-end QA","2026-08-11","production","Ask Magic Mike /sell","Suppressed","One delivered internal alert","Audited","Reconciled","Retain as QA evidence"],
        ["bbed9a2d-4619-4c18-9298-5167a9694f73","Production cutover home-value QA","2026-08-11","production","Ask Magic Mike /value","Suppressed","One sent internal alert","Audited","Reconciled","Retain as QA evidence"],
        ["70f63f35-2478-4738-b84c-bc1a89b8482c","WordPress Form 3 bridge acceptance","2026-08-14","production","Gravity Form 3","Suppressed","One delivered internal alert","Audited","Reconciled","Retain as QA evidence"],
        ["a7b1cf10-e546-48c4-85b1-2dee424ab156","Pre-fix idempotency replay","2026-08-14","production","Gravity Form 3 replay","Suppressed after guarded transaction","None","Suppression audit added","Reconciled","Retain as incident evidence"],
    ]
    title_sheet(r,"Production Tests","Aggregate at audit: 6 tests, 0 live Neon prospects, 0 unsuppressed tests.",len(headers)); write_table(r,headers,rows,widths=[39,42,14,16,30,34,34,24,24,30])
    c=wb.create_sheet("Automated Checks"); title_sheet(c,"Automated Checks","Required assertions for each release and daily reconciliation.",4); write_table(c,["Check","Expected","Current result","Evidence"],[["Unsuppressed tests",0,0,"Neon aggregate query"],["Live lead marked test",0,"Requires first-live-lead reconciliation","FIRST_LIVE_LEAD_RESPONSE_RUNBOOK.md"],["Tests in executive KPIs",0,0,"Scoreboard COUNTIFS excludes Test?=Yes"],["Test consumer messages",0,0,"Lead alert service suppression"],["Production tests with real customer data",0,0,"QA evidence review"]],widths=[38,20,34,50])
    return save(wb,"TEST_LEAD_REGISTER.xlsx")


def build_budget_roi():
    wb,ws=base_book("Budget and ROI Workbook","Free-first planning model. All business outcomes are assumptions until real data is entered.")
    a=wb.create_sheet("Assumptions"); title_sheet(a,"Assumptions","Yellow cells are editable assumptions; current paid media spend is zero.",5)
    rows=[["Monthly paid media budget",0,0,0,"USD"],["Monthly organic leads",5,12,25,"leads"],["Qualified rate",0.25,0.35,0.45,"%"],["Appointment rate from qualified",0.25,0.35,0.45,"%"],["Signed-client rate from appointments",0.2,0.3,0.4,"%"],["Closing rate from signed clients",0.5,0.65,0.75,"%"],["Gross revenue per closing",5000,7500,10000,"USD"],["Monthly operating tools",0,0,0,"USD"]]
    write_table(a,["Input","Conservative","Base","Upside","Unit"],rows,start_row=4,widths=[42,18,18,18,16])
    for row in range(5,13):
        for col in range(2,5): a.cell(row,col).fill=PatternFill("solid",fgColor=INPUT)
    for row in [7,8,9,10]:
        for col in range(2,5): a.cell(row,col).number_format="0%"
    for row in [5,11,12]:
        for col in range(2,5): a.cell(row,col).number_format='$#,##0'
    m=wb.create_sheet("ROI Model"); title_sheet(m,"ROI Model","Formula-driven scenarios; tests excluded because assumptions represent genuine pipeline only.",5)
    m.append([]); m.append([]); m.append([])
    headers=["Metric","Conservative","Base","Upside","Formula note"]; write_table(m,headers,[],start_row=4,widths=[38,18,18,18,48])
    metrics=[("Organic leads","='Assumptions'!B6","='Assumptions'!C6","='Assumptions'!D6","Input"),("Qualified leads","=B5*'Assumptions'!B7","=C5*'Assumptions'!C7","=D5*'Assumptions'!D7","Leads x qualified rate"),("Appointments","=B6*'Assumptions'!B8","=C6*'Assumptions'!C8","=D6*'Assumptions'!D8","Qualified x appointment rate"),("Signed clients","=B7*'Assumptions'!B9","=C7*'Assumptions'!C9","=D7*'Assumptions'!D9","Appointments x signed rate"),("Closings","=B8*'Assumptions'!B10","=C8*'Assumptions'!C10","=D8*'Assumptions'!D10","Signed x closing rate"),("Gross attributed revenue","=B9*'Assumptions'!B11","=C9*'Assumptions'!C11","=D9*'Assumptions'!D11","Closings x gross revenue"),("Total modeled cost","='Assumptions'!B5+'Assumptions'!B12","='Assumptions'!C5+'Assumptions'!C12","='Assumptions'!D5+'Assumptions'!D12","Paid budget + tools"),("Modeled ROI","=IF(B11=0,0,(B10-B11)/B11)","=IF(C11=0,0,(C10-C11)/C11)","=IF(D11=0,0,(D10-D11)/D11)","(Revenue - cost) / cost; zero when cost is zero")]
    for idx,(label,b,c,d,note) in enumerate(metrics,5):
        m.cell(idx,1,label); m.cell(idx,2,b); m.cell(idx,3,c); m.cell(idx,4,d); m.cell(idx,5,note)
        for col in range(1,6): m.cell(idx,col).border=Border(bottom=THIN)
    for row in [10,11]:
        for col in range(2,5): m.cell(row,col).number_format='$#,##0'
    for col in range(2,5): m.cell(12,col).number_format='0%'
    return save(wb,"BUDGET_AND_ROI_WORKBOOK.xlsx")


def build_evidence_register():
    wb,ws=base_book("Evidence and Assumption Register","Separates verified production facts from hypotheses and owner decisions.")
    e=wb.create_sheet("Evidence Register"); headers=["ID","Statement","Classification","Status","Source","Verified date","Owner","Next review","Notes"]
    rows=[
        ["E-001","AskMagicMike.com serves the canonical app","Evidence","VERIFIED","Vercel deployment dpl_3ogimm1EhHCaPkEfXLAeojrm2H8Z","2026-08-15","System owner","After next deploy","Production commit e754456cecaf6538df25bb4bf5eebe57ebf6eacb; apex redirects to www"],
        ["E-002","Form 3 creates one canonical lead and one internal alert","Evidence","VERIFIED","docs/FORM3_PRODUCTION_RECONCILIATION.md","2026-08-14","System owner","Daily","Replay idempotent"],
        ["E-003","Neon contains six suppressed QA leads and zero genuine live prospects at audit","Evidence","POINT-IN-TIME","Neon production aggregate at 2026-08-15T16:35:23Z","2026-08-15","System owner","Daily","Zero QA in Active/New; zero unsuppressed QA"],
        ["E-004","Web Push has zero active devices","Evidence","VERIFIED","Neon staff_push_subscriptions aggregate","2026-08-15","System owner","After physical enrollment","Email remains primary"],
        ["E-005","Meta crawler receives 403 on two Our Town URLs only","Evidence","OPEN HOSTING ACTION","40/42 crawler matrix; other crawlers and Ask Magic Mike URLs pass","2026-08-15","Hosting operator","After narrowly scoped host review","Do not weaken the firewall broadly"],
        ["E-006","First-live and SLA monitors are scheduled","Evidence","VERIFIED","Vercel cron logs and vercel.json","2026-08-15","System owner","Daily","First-live every 2 minutes; SLA hourly"],
        ["A-001","Organic leads may reach 12/month","Assumption","UNVALIDATED","Budget model base case","2026-08-14","Brokerage owner","After 30 days","Do not present as forecasted fact"],
        ["D-001","Form 7 consent wording","Owner decision","PENDING","OWNER_ACTIONS_REMAINING.md","2026-08-14","Brokerage owner/BIC","Before Form 7 activation","No Constant Contact activation before approval"],
    ]
    title_sheet(e,"Evidence Register","Do not convert assumptions into claims.",len(headers)); write_table(e,headers,rows,widths=[12,60,24,24,58,18,24,20,50]); add_status_rules(e,"D",5,12)
    return save(wb,"EVIDENCE_AND_ASSUMPTION_REGISTER.xlsx")


builders=[build_activation_matrix,build_scoreboard,build_roster,build_access_matrix,build_push_register,build_monitoring_matrix,build_attribution_matrix,build_test_register,build_budget_roi,build_evidence_register]

if __name__ == "__main__":
    paths=[str(builder()) for builder in builders]
    print("\n".join(paths))
